from openpyxl import Workbook, load_workbook

wb = load_workbook('Grades.xlsx')
ws = wb.active

ws.merge_cells("A1:D1")
ws.unmerge_cells("A1:D1")

wb.save('Grades.xlsx')