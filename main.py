from openpyxl import Workbook, load_workbook

wb = load_workbook('Grades.xlsx')
ws = wb.active
print(ws["A1"].value)
ws["A1"] = "Name1"

wb.create_sheet("Test")

print(wb.sheetnames)


# wb.save("Grades.xlsx")
