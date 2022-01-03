from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook("Grades.xlsx")
ws = wb.active

for row in range(1,11):
    for col in range(1, 5):
        char = get_column_letter(col)  # char = chr(65+col) (0,-)
        ws[char + str(row)] = char + str(row)

wb.save("Grades.xlsx")